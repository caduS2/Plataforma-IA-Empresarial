from __future__ import annotations

from io import BytesIO
from pathlib import Path

from docx import Document as WordDocument
from openpyxl import load_workbook
from pypdf import PdfReader


FORMATOS_SUPORTADOS = {
    ".txt": {"text/plain"}, ".md": {"text/markdown", "text/plain"}, ".csv": {"text/csv", "application/csv", "text/plain"},
    ".pdf": {"application/pdf"}, ".docx": {"application/vnd.openxmlformats-officedocument.wordprocessingml.document"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}, ".xlsm": {"application/vnd.ms-excel.sheet.macroenabled.12"},
    ".png": {"image/png"}, ".jpg": {"image/jpeg"}, ".jpeg": {"image/jpeg"},
}


def validar_arquivo(nome_arquivo: str, tipo_mime: str | None, conteudo: bytes) -> str:
    suffix = Path(nome_arquivo).suffix.lower()
    if suffix not in FORMATOS_SUPORTADOS:
        raise ValueError("Formato de arquivo não permitido.")
    if tipo_mime and tipo_mime not in FORMATOS_SUPORTADOS[suffix]:
        raise ValueError("O tipo informado não corresponde ao formato permitido.")
    if suffix == ".pdf" and not conteudo.startswith(b"%PDF-"):
        raise ValueError("O PDF é inválido.")
    if suffix == ".png" and not conteudo.startswith(b"\x89PNG\r\n\x1a\n"):
        raise ValueError("A imagem PNG é inválida.")
    if suffix in {".jpg", ".jpeg"} and not conteudo.startswith(b"\xff\xd8\xff"):
        raise ValueError("A imagem JPEG é inválida.")
    if suffix in {".docx", ".xlsx", ".xlsm"} and not conteudo.startswith(b"PK\x03\x04"):
        raise ValueError("O arquivo Office é inválido.")
    return suffix


def extrair_texto(nome_arquivo: str, conteudo: bytes) -> str | None:
    """Extracts readable text from the supported office formats.

    The raw file remains the source of record; extraction is used only for
    retrieval and source citations.
    """
    suffix = Path(nome_arquivo).suffix.lower()
    try:
        if suffix in {".txt", ".md", ".csv"}:
            return conteudo.decode("utf-8", errors="replace")[:500_000]
        if suffix == ".pdf":
            return "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(conteudo)).pages)[:500_000]
        if suffix == ".docx":
            document = WordDocument(BytesIO(conteudo))
            return "\n".join(paragraph.text for paragraph in document.paragraphs)[:500_000]
        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in workbook.worksheets:
                lines.append(f"Planilha: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        lines.append(" | ".join(values))
            return "\n".join(lines)[:500_000]
    except Exception:
        return None
    return None
